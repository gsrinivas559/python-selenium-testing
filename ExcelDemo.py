import openpyxl

book = openpyxl.load_workbook("ExcelDemo.xlsx")
sheet = book.active
# Reading cell value
cell = sheet.cell(row=1, column=2)
print(cell.value)  # FirstName

# Writing cell value
sheet.cell(row=2, column=2).value = "Srinivas"
print(sheet.cell(row=2, column=2).value)  # Srinivas

# Max rows in a sheet
print(sheet.max_row)  # 10

# Max columns in a sheet
print(sheet.max_column)  # 4

# Another way of getting value from excel cell is giving cell name.i.e, "A5, B2"
print(sheet["B2"].value)  # Srinivas

# Reading all the values from sheet
print("********* Reading all the cell values **********")
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=" ")

    print()  # to move to a new line after completing each column values

# Reading all the column values for matching row from sheet
print("********* Reading column values for matching row value **********")
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Resource2":
        for j in range(2, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value, end=" ")  # AutoUser2 Test2 TestUser2
        print()  # to move to a new line after completing each column values

# Reading all the column values for matching row and storing the data in dictionary from sheet
print("********* Reading column values for matching row value and storing the data in dictionary **********")
Dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Resource2":
        for j in range(2, sheet.max_column + 1):
            print(sheet.cell(row=1, column=j).value + "=" + sheet.cell(row=i, column=j).value)  # FirstName=AutoUser2
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)  # {'FirstName': 'AutoUser2', 'LastName': 'Test2', 'Password': 'TestUser2'}
