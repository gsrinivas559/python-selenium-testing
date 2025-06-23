from time import sleep

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


driver = webdriver.Chrome()  # chrome driver initiation
driver.implicitly_wait(5)  # globally wait applied
driver.maximize_window()
driver.get("https://rahulshettyacademy.com/upload-download-test/")  # testing rahul shetty practice page

# Scenario- Download the file, update the data, re-upload the file and validate data changed
file_path = "C:\\Users\\gadwa\\Downloads\\download.xlsx"
fruit_name = "Apple"
newValue = "990"
# download
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()
sleep(5)  # To visually view the output

# edit the Excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

# upload
driver.find_element(By.CSS_SELECTOR, "#fileinput").send_keys(file_path)
WebDriverWait(driver, 10).until(
    expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body")))
alertMsg = driver.find_element(By.CSS_SELECTOR, ".Toastify__toast-body").text
assert alertMsg == "Updated Excel Data Successfully."
sleep(2)  # To visually view the output

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
fruitPrice = "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']"
actual_price = driver.find_element(By.XPATH, fruitPrice).text
assert actual_price == newValue
